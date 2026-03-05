"""
Stage 1 + 2: Excel Ingestion Pipeline
Reads raw Excel workbooks into the raw layer, populates curated tables,
and generates i18n keys for translatable fields.

Uses openpyxl directly (lightweight) instead of pandas (heavy ~150MB).
"""
import asyncio
import hashlib
import json
import math
import os
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# Ensure backend is on path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import BASE_DIR, DB_PATH, DATA_DIR
from db.database import init_db, execute_query, execute_write

# ── File paths ──────────────────────────────────────────────
PRODUCTS_FILE = BASE_DIR / "products-export.xlsx"
ORDERS_FILE   = BASE_DIR / "Consumer Order History 1.xlsx"

# ── Helpers ─────────────────────────────────────────────────

def _is_empty(val: Any) -> bool:
    """Check if a value is None or NaN."""
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return False


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _cell_type(val: Any) -> str:
    if _is_empty(val):
        return "empty"
    if isinstance(val, bool):
        return "boolean"
    if isinstance(val, int):
        return "integer"
    if isinstance(val, float):
        return "number"
    if isinstance(val, datetime):
        return "datetime"
    return "string"


def _cell_values(val: Any) -> Dict[str, Any]:
    """Return typed value columns for sheet_cells."""
    vtype = _cell_type(val)
    out: Dict[str, Any] = {
        "value_type": vtype,
        "value_text": None,
        "value_number": None,
        "value_boolean": None,
        "value_datetime": None,
    }
    if vtype == "empty":
        pass
    elif vtype == "boolean":
        out["value_boolean"] = val
        out["value_text"] = str(val)
    elif vtype in ("integer", "number"):
        out["value_number"] = float(val)
        out["value_text"] = str(val)
    elif vtype == "datetime":
        out["value_datetime"] = val.isoformat()
        out["value_text"] = val.isoformat()
    else:
        out["value_text"] = str(val)
    return out


def _safe(val: Any) -> Any:
    """Make a value JSON-safe."""
    if _is_empty(val):
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return val


def _i18n_key(namespace: str, entity_id: Any, field: str) -> str:
    """Generate a stable i18n key, e.g. product.16066.name"""
    return f"{namespace}.{entity_id}.{field}"


def _read_sheet_rows(file_path: Path, sheet_name: str) -> List[List[Any]]:
    """Read all rows from an Excel sheet using openpyxl. Returns list of row-lists."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows():
        rows.append([cell.value for cell in row])
    wb.close()
    return rows


# ── Stage 1: Raw ingestion ──────────────────────────────────

async def _create_import_run(notes: str = "") -> int:
    run_uuid = str(uuid.uuid4())
    return await execute_write(
        "INSERT INTO import_runs(run_uuid, source_system, notes) VALUES (?, ?, ?)",
        (run_uuid, "excel", notes),
    )


async def _register_workbook(
    run_id: int, file_path: Path, source_language: str = "de"
) -> int:
    stat = file_path.stat()
    return await execute_write(
        """INSERT INTO source_workbooks
           (import_run_id, file_name, file_path, file_size_bytes,
            file_hash_sha256, source_language, row_count_estimate)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (
            run_id,
            file_path.name,
            str(file_path),
            stat.st_size,
            _sha256(file_path),
            source_language,
            None,  # filled later
        ),
    )


async def _register_sheet(
    workbook_id: int, sheet_name: str, sheet_index: int,
    max_row: int, max_col: int, header_row: Optional[int]
) -> int:
    return await execute_write(
        """INSERT INTO workbook_sheets
           (workbook_id, sheet_name, sheet_index, max_row, max_column, header_row_index)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (workbook_id, sheet_name, sheet_index, max_row, max_col, header_row),
    )


async def _insert_row(sheet_id: int, row_idx: int, kind: str, values: list) -> int:
    return await execute_write(
        "INSERT INTO sheet_rows(sheet_id, row_index, row_kind, raw_values_json) VALUES (?, ?, ?, ?)",
        (sheet_id, row_idx, kind, json.dumps([_safe(v) for v in values], ensure_ascii=False)),
    )


async def _insert_cell(
    sheet_id: int, row_idx: int, col_idx: int,
    col_label: str, header_name: Optional[str], val: Any
) -> int:
    cv = _cell_values(val)
    return await execute_write(
        """INSERT INTO sheet_cells
           (sheet_id, row_index, col_index, col_label, header_name,
            value_type, value_text, value_number, value_boolean, value_datetime)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            sheet_id, row_idx, col_idx, col_label, header_name,
            cv["value_type"], cv["value_text"], cv["value_number"],
            cv["value_boolean"], cv["value_datetime"],
        ),
    )


async def _log_issue(
    run_id: int, wb_id: int, sh_id: int,
    row: Optional[int], col: Optional[int],
    severity: str, code: str, message: str,
    context: Optional[dict] = None,
):
    await execute_write(
        """INSERT INTO ingestion_issues
           (import_run_id, workbook_id, sheet_id, row_index, col_index,
            severity, issue_code, issue_message, context_json)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (run_id, wb_id, sh_id, row, col, severity, code, message,
         json.dumps(context, ensure_ascii=False) if context else None),
    )


# ── Stage 2: i18n key registration ──────────────────────────

async def _register_i18n_key(
    namespace: str, string_key: str,
    source_language: str, source_text: str
) -> Optional[int]:
    """Insert a localized_string row if it doesn't already exist. Return its id."""
    if not source_text or not source_text.strip():
        return None
    existing = await execute_query(
        "SELECT id FROM localized_strings WHERE namespace=? AND string_key=?",
        (namespace, string_key),
    )
    if existing:
        return existing[0]["id"]
    return await execute_write(
        """INSERT INTO localized_strings(namespace, string_key, source_language, source_text)
           VALUES (?, ?, ?, ?)""",
        (namespace, string_key, source_language, source_text),
    )


# ── Products ingestion ──────────────────────────────────────

async def ingest_products(run_id: int) -> Tuple[int, int]:
    """Ingest products-export.xlsx. Returns (workbook_id, row_count)."""
    if not PRODUCTS_FILE.exists():
        print(f"  ⚠ Products file not found: {PRODUCTS_FILE}")
        return 0, 0

    wb_id = await _register_workbook(run_id, PRODUCTS_FILE, "de")
    all_rows = _read_sheet_rows(PRODUCTS_FILE, "Products")

    if not all_rows:
        return wb_id, 0

    # First row is the header
    headers = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(all_rows[0])]
    data_rows = all_rows[1:]
    nrows = len(data_rows)
    ncols = len(headers)

    sh_id = await _register_sheet(wb_id, "Products", 0, nrows, ncols, 0)

    # Raw header row
    await _insert_row(sh_id, 0, "header", headers)
    for ci, h in enumerate(headers):
        await _insert_cell(sh_id, 0, ci, chr(65 + ci) if ci < 26 else f"C{ci}", None, h)

    # Build a header-name-to-index map (lowercase for flexible lookup)
    hdr_map = {str(h).lower(): i for i, h in enumerate(headers)}

    inserted = 0
    for ri, vals in enumerate(data_rows):
        row_idx = ri + 1  # 0 is header

        # Pad vals to match header length
        while len(vals) < ncols:
            vals.append(None)

        # Raw layer
        await _insert_row(sh_id, row_idx, "data", vals)
        for ci, val in enumerate(vals):
            await _insert_cell(
                sh_id, row_idx, ci,
                chr(65 + ci) if ci < 26 else f"C{ci}",
                str(headers[ci]) if ci < len(headers) else None, val,
            )

        # Curated layer — look up columns by header name
        def _col(name: str, default=None):
            idx = hdr_map.get(name.lower())
            if idx is not None and idx < len(vals) and not _is_empty(vals[idx]):
                return vals[idx]
            return default

        pid = int(_col("product id", 0))
        pname = str(_col("product name", ""))
        pzn_raw = _col("pzn")
        pzn = int(pzn_raw) if pzn_raw is not None else None
        price_raw = _col("price rec")
        price = float(price_raw) if price_raw is not None else None
        pkg = str(_col("package size", "")) if _col("package size") is not None else None
        desc = str(_col("descriptions", "")) if _col("descriptions") is not None else None

        # i18n keys
        name_key = _i18n_key("product", pid, "name")
        desc_key = _i18n_key("product", pid, "description")

        await _register_i18n_key("product", name_key, "de", pname)
        if desc:
            await _register_i18n_key("product", desc_key, "de", desc)

        raw_json = json.dumps({str(headers[i]): _safe(v) for i, v in enumerate(vals) if i < len(headers)}, ensure_ascii=False)

        try:
            await execute_write(
                """INSERT INTO products_export_records
                   (workbook_id, sheet_id, source_row_index, source_language,
                    product_id, product_name, product_name_i18n_key,
                    pzn, price_rec_eur, package_size,
                    descriptions, descriptions_i18n_key,
                    translation_status, raw_record_json)
                   VALUES (?, ?, ?, 'de', ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)""",
                (wb_id, sh_id, row_idx, pid, pname, name_key,
                 pzn, price, pkg, desc, desc_key, raw_json),
            )
            inserted += 1
        except Exception as e:
            await _log_issue(run_id, wb_id, sh_id, row_idx, None,
                             "error", "PRODUCT_INSERT", str(e),
                             {"product_id": pid, "product_name": pname})

    # Update row count estimate
    await execute_write(
        "UPDATE source_workbooks SET row_count_estimate = ? WHERE id = ?",
        (nrows, wb_id),
    )

    print(f"  ✅ Products: {inserted}/{nrows} rows ingested")
    return wb_id, inserted


# ── Consumer Order History ingestion ─────────────────────────

async def ingest_orders(run_id: int) -> Tuple[int, int]:
    """Ingest Consumer Order History 1.xlsx. Returns (workbook_id, row_count)."""
    if not ORDERS_FILE.exists():
        print(f"  ⚠ Orders file not found: {ORDERS_FILE}")
        return 0, 0

    wb_id = await _register_workbook(run_id, ORDERS_FILE, "de")

    # Read raw — all rows as-is
    all_rows = _read_sheet_rows(ORDERS_FILE, "Sheet1")
    nrows = len(all_rows)
    ncols = max(len(r) for r in all_rows) if all_rows else 0

    # Detect layout: row 0 = title, row 2 = subtitle, row 4 = header, rows 5+ = data
    TITLE_ROW = 0
    SUBTITLE_ROW = 2
    HEADER_ROW = 4
    DATA_START = 5

    header_vals = all_rows[HEADER_ROW] if HEADER_ROW < nrows else []
    headers = [str(v) if not _is_empty(v) else f"col_{i}" for i, v in enumerate(header_vals)]

    sh_id = await _register_sheet(wb_id, "Sheet1", 0, nrows, ncols, HEADER_ROW)

    # Store sheet metadata
    title_text = str(all_rows[TITLE_ROW][0]) if TITLE_ROW < nrows and all_rows[TITLE_ROW] and not _is_empty(all_rows[TITLE_ROW][0]) else None
    subtitle_text = str(all_rows[SUBTITLE_ROW][0]) if SUBTITLE_ROW < nrows and all_rows[SUBTITLE_ROW] and not _is_empty(all_rows[SUBTITLE_ROW][0]) else None
    await execute_write(
        """INSERT INTO consumer_order_history_sheet_metadata
           (workbook_id, sheet_id, title_text, subtitle_text, header_row_index, header_json)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (wb_id, sh_id, title_text, subtitle_text, HEADER_ROW,
         json.dumps(headers, ensure_ascii=False)),
    )

    # Raw: ingest ALL rows (title, blank, header, data)
    for ri in range(nrows):
        vals = all_rows[ri]
        # Pad to ncols
        while len(vals) < ncols:
            vals.append(None)

        non_null = [v for v in vals if not _is_empty(v)]
        if ri == TITLE_ROW:
            kind = "title"
        elif ri == SUBTITLE_ROW:
            kind = "subtitle"
        elif ri == HEADER_ROW:
            kind = "header"
        elif len(non_null) == 0:
            kind = "blank"
        else:
            kind = "data"

        await _insert_row(sh_id, ri, kind, vals)
        for ci, val in enumerate(vals):
            hdr = headers[ci] if ci < len(headers) and ri >= DATA_START else None
            await _insert_cell(sh_id, ri, ci, chr(65 + ci) if ci < 26 else f"C{ci}", hdr, val)

    # Curated: data rows only
    inserted = 0
    for ri in range(DATA_START, nrows):
        vals = all_rows[ri]
        # Pad to ncols
        while len(vals) < ncols:
            vals.append(None)

        non_null = [v for v in vals if not _is_empty(v)]
        if len(non_null) == 0:
            continue  # skip blank rows

        def _get(col_idx: int) -> Any:
            return vals[col_idx] if col_idx < len(vals) and not _is_empty(vals[col_idx]) else None

        patient_id = str(_get(0)) if _get(0) else None
        if not patient_id:
            await _log_issue(run_id, wb_id, sh_id, ri, 0,
                             "warning", "MISSING_PATIENT_ID", "Row has no patient_id")
            continue

        patient_age = int(_get(1)) if _get(1) is not None else None
        patient_gender = str(_get(2)) if _get(2) else None
        purchase_date_raw = _get(3)
        purchase_date = None
        if purchase_date_raw is not None:
            if isinstance(purchase_date_raw, datetime):
                purchase_date = purchase_date_raw.strftime("%Y-%m-%d")
            else:
                purchase_date = str(purchase_date_raw)

        product_name = str(_get(4)) if _get(4) else ""
        quantity = int(_get(5)) if _get(5) is not None else None
        total_price = float(_get(6)) if _get(6) is not None else None
        dosage_freq = str(_get(7)) if _get(7) else None
        rx_required = str(_get(8)) if _get(8) else None

        # Normalize known fields deterministically
        gender_norm = None
        if patient_gender:
            g = patient_gender.strip().upper()
            gender_norm = {"M": "male", "F": "female", "D": "diverse"}.get(g, patient_gender.lower())

        freq_norm = None
        if dosage_freq:
            freq_norm = _normalize_frequency(dosage_freq)

        rx_bool = None
        if rx_required:
            rx_bool = _normalize_prescription(rx_required)

        # i18n key for product name
        order_key = f"r{ri}"
        name_key = _i18n_key("order_product", order_key, "name")
        await _register_i18n_key("order_product", name_key, "de", product_name)

        raw_json = json.dumps(
            {headers[i]: _safe(v) for i, v in enumerate(vals) if i < len(headers)},
            ensure_ascii=False,
        )

        try:
            await execute_write(
                """INSERT INTO consumer_order_history_records
                   (workbook_id, sheet_id, source_row_index, source_language,
                    patient_id, patient_age, patient_gender, patient_gender_norm,
                    purchase_date, product_name, product_name_i18n_key,
                    quantity, total_price_eur, dosage_frequency, dosage_frequency_norm,
                    prescription_required, prescription_required_bool,
                    translation_status, raw_record_json)
                   VALUES (?, ?, ?, 'de', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)""",
                (wb_id, sh_id, ri,
                 patient_id, patient_age, patient_gender, gender_norm,
                 purchase_date, product_name, name_key,
                 quantity, total_price, dosage_freq, freq_norm,
                 rx_required, rx_bool, raw_json),
            )
            inserted += 1
        except Exception as e:
            await _log_issue(run_id, wb_id, sh_id, ri, None,
                             "error", "ORDER_INSERT", str(e),
                             {"patient_id": patient_id})

    await execute_write(
        "UPDATE source_workbooks SET row_count_estimate = ? WHERE id = ?",
        (nrows, wb_id),
    )

    print(f"  ✅ Orders: {inserted}/{nrows - DATA_START} data rows ingested")
    return wb_id, inserted


# ── Deterministic normalizers (no LLM) ─────────────────────

_FREQ_MAP = {
    # English (already in data)
    "once daily": "once_daily",
    "twice daily": "twice_daily",
    "three times daily": "three_times_daily",
    "as needed": "as_needed",
    # German equivalents
    "einmal täglich": "once_daily",
    "zweimal täglich": "twice_daily",
    "dreimal täglich": "three_times_daily",
    "bei bedarf": "as_needed",
    "nach bedarf": "as_needed",
    "morgens": "once_daily_morning",
    "abends": "once_daily_evening",
}

_RX_MAP = {
    "yes": True, "no": False,
    "ja": True, "nein": False,
    "true": True, "false": False,
    "1": True, "0": False,
}


def _normalize_frequency(raw: str) -> Optional[str]:
    return _FREQ_MAP.get(raw.strip().lower())


def _normalize_prescription(raw: str) -> Optional[bool]:
    return _RX_MAP.get(raw.strip().lower())


# ── Seed default languages ──────────────────────────────────

async def seed_languages():
    """Insert default supported languages."""
    langs = [
        ("de", "German", "Deutsch", 1, 1),
        ("en", "English", "English", 1, 0),
        ("hi", "Hindi", "हिन्दी", 1, 0),
        ("fr", "French", "Français", 1, 0),
        ("es", "Spanish", "Español", 1, 0),
        ("ar", "Arabic", "العربية", 1, 0),
        ("tr", "Turkish", "Türkçe", 1, 0),
    ]
    for code, name, native, active, default in langs:
        try:
            await execute_write(
                "INSERT OR IGNORE INTO languages(code, name, native_name, is_active, is_default) VALUES (?, ?, ?, ?, ?)",
                (code, name, native, active, default),
            )
        except Exception:
            pass
    print("  ✅ Languages seeded")


# ── Main entry point ────────────────────────────────────────

async def run_ingestion() -> int:
    """Run full Stage 1+2 ingestion. Returns the import_run id."""
    print("━━ Stage 1+2: Excel Ingestion ━━━━━━━━━━━━━━━━━━━━━")
    await init_db()
    await seed_languages()

    run_id = await _create_import_run("Initial Excel import")

    try:
        _, p_count = await ingest_products(run_id)
        _, o_count = await ingest_orders(run_id)

        status = "success" if (p_count > 0 and o_count > 0) else "partial"
        await execute_write(
            "UPDATE import_runs SET status=?, finished_at=CURRENT_TIMESTAMP WHERE id=?",
            (status, run_id),
        )
        print(f"  ✅ Import run {run_id} finished ({status})")
    except Exception as e:
        await execute_write(
            "UPDATE import_runs SET status='failed', finished_at=CURRENT_TIMESTAMP WHERE id=?",
            (run_id,),
        )
        print(f"  ❌ Import run {run_id} failed: {e}")
        raise

    return run_id


if __name__ == "__main__":
    asyncio.run(run_ingestion())
